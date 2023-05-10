from urllib.request import urlopen,Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font

# scrape the website below to retrieve the top 5 countries with the highest GDPs. Calculate the GDP per capita
# by dividing the GDP by the population. You can perform the calculation in Python natively or insert the code
# in excel that will perform the calculation in Excel by each row. DO NOT scrape the GDP per capita from the
# webpage, make sure you use your own calculation.


### REMEMBER ##### - your output should match the excel file (GDP_Report.xlsx) including all formatting.

webpage = 'https://www.worldometers.info/gdp/gdp-by-country/'

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(webpage, headers=headers)
webpage = urlopen(req).read()			
soup = BeautifulSoup(webpage, 'html.parser')
print(soup.title.text)


tables = soup.findAll('table')
updated_tables = tables[0]
rows = updated_tables.findAll('tr')

wb = xl.Workbook()
ws = wb.active
ws.title = "GDP Report"

ws['A1'] = "No."
ws['B1'] = "Country"
ws['C1'] = "GDP"
ws['D1'] = "Population"
ws['E1'] = "GDP Per Capita"

header_font = Font(name= 'Times New Roman', size=16, bold=True)

ws['A1'].font = header_font
ws['B1'].font = header_font
ws['C1'].font = header_font
ws['D1'].font = header_font
ws['E1'].font = header_font

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 28

for x in range(1, 6):
    td = rows[x].findAll('td')
    no = td[0].text
    country = td[1].text
    gdp = float(td[2].text.replace('$', "").replace(',', ''))
    population = int(td[5].text.replace(',', ''))
    gdp_per_capita = round(gdp / population, 2)

    ws['A' + str(x+1)] = no
    ws['B' + str(x+1)] = country
    ws['C' + str(x+1)] = '$' + str(format(gdp, ',.0f'))
    ws['D' + str(x+1)] = str(format(population, ','))
    ws['E' + str(x+1)] = '$' + str(format(gdp_per_capita, ',.2f'))


wb.save("GDP Report.xlsx")
    




    





