import openpyxl as xl
from openpyxl.styles import Font

wb = xl.Workbook()

ws = wb.active

ws.title = 'First Sheet'

wb.create_sheet(index=1, title='Second Sheet')


ws['A1'] = 'Invoice'

ws['A1'].font = Font(name='Times New Roman', size=24, bold=True, italic=False)

myfont = Font(name='Times New Roman', size=24, bold=True, italic=False)

ws['A1'].font = myfont

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws.merge_cells('A1:B1')

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

ws['A8'] = 'Total'

ws['A8'].font = myfont

ws['B8'] = '=SUM(B2:B4)'


write_sheet = wb['Second Sheet']


read_wb = xl.load_workbook('ProduceReport.xlsx')
read_ws = read_wb['ProduceReport']

maxC = read_ws.max_column
maxR = read_ws.max_row


write_sheet['A1'] = 'Produce'
write_sheet['B1'] = 'Cost Per Pound'
write_sheet['C1'] = 'Amt Sold'
write_sheet['D1'] = 'Total'

write_row = 2
write_colA = 1
write_colB = 2
write_colC = 3
write_colD = 4


for currentrow in read_ws.iter_rows(min_row=2, max_row=maxR, max_col=maxC):
    name = currentrow[0].value
    cost = float(currentrow[1].value)
    amt_sold = float(currentrow[2].value)
    total = float(currentrow[3].value)

    write_sheet.cell(write_row, write_colA).value = name     #order is row, column
    write_sheet.cell(write_row, write_colB).value = cost
    write_sheet.cell(write_row, write_colC).value = amt_sold
    write_sheet.cell(write_row, write_colD).value = total

    write_row += 1

summary_row = write_row + 1

write_sheet['B' + str(summary_row)] = 'Total'
    






wb.save('PythontoExcel.xlsx')





