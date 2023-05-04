from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
import keys
from twilio.rest import Client


url = 'https://www.coingecko.com/'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(url, headers=headers)
webpage = urlopen(req).read()
soup = BeautifulSoup(webpage, 'html.parser')
print(soup.title.text)


client = Client(keys.accountSID, keys.authToken)

#TwilioNumber = "+15075167290"

#mycellphone = "+18322312653"

TwilioNumber = ""

mycellphone = ""


tables = soup.findAll('table')
updated_tables = tables[0]
rows = updated_tables.findAll('tr')

wb = xl.Workbook()
ws = wb.active
ws.title = "Cryptocurrencies"

ws['A1'] = "#"
ws['B1'] = "Cryptocurrency"
ws['C1'] = "Price"
ws['D1'] = "Percent change within 24 hours"
ws['E1'] = "Price based on change"

header_font = Font(name= 'Times New Roman', size=14, bold=True, underline='single')

ws['A1'].font = header_font
ws['B1'].font = header_font
ws['C1'].font = header_font
ws['D1'].font = header_font
ws['E1'].font = header_font

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 30

for cell in ws[2:2]:
    cell.font = Font(name="Times New Roman")
for cell in ws[3:3]:
    cell.font = Font(name="Times New Roman")
for cell in ws[4:4]:
    cell.font = Font(name="Times New Roman")
for cell in ws[5:5]:
    cell.font = Font(name="Times New Roman")
for cell in ws[6:6]:
    cell.font = Font(name="Times New Roman")


for row in range(1, 6):
    td = rows[row].findAll('td')
    number = td[1].text
    cryptocurrency = td[2].text + ""
    price = float(td[3].text.replace(",", "").replace("$", ""))
    changed_percent = float(td[5].text.replace("%", ""))
    total_change = round((price + changed_percent), 2)
    new_price = int(total_change - price)
    if new_price <= -5 or new_price >= 5:
        text = client.messages.create(to=mycellphone, from_=TwilioNumber, body="A change of $5 has occurred")
        print(text.status)


    ws['A' + str(row+1)] = number
    ws['B' + str(row+1)] = cryptocurrency
    ws['C' + str(row+1)] = '$' + str(format(price, ',.2f'))
    ws['D' + str(row+1)] = str(format(changed_percent, ',.2f') + '%') 
    ws['E' + str(row+1)] = '$' + format(total_change, ',.2f')



wb.save("Cryptocurrencies.xlsx")